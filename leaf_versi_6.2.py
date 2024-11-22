#PROGRAM FINAL OK
#Program Update
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import QFileDialog, QLabel
import cv2, imutils
from ultralytics import YOLO
import cvzone
import pyshine as ps
from PIL import Image
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QWidget, QCheckBox, QLabel, QVBoxLayout, QMessageBox, QFileDialog, QMainWindow, QDialog

from PyQt5.QtGui import QImage
from PyQt5.QtGui import QPixmap
#library sensor jarak
import adafruit_vl53l0x
import time
import board
import busio
#library untuk kamera dan yolo
from ultralytics import YOLO
import cv2
import numpy as np
import math
import sys
import imutils
import shutil
from shutil import copy
import warnings
#library tambahan program lama
from PyQt5 import QtWidgets, uic
from hashlib import new
from PyQt5 import uic
from PyQt5.QtWidgets import *
from PyQt5.QtMultimedia import QCameraInfo
from PyQt5.QtWidgets import QApplication, QMainWindow, QDialog, QMessageBox, QFileDialog
from PyQt5.QtCore import QThread, pyqtSignal, pyqtSlot, QTimer, QDateTime, Qt, QTime
from PyQt5.QtGui import QImage, QPixmap, QPainter, QPen, QCursor
import sysinfo
import random as rnd
import os
import openpyxl
from openpyxl.styles import Alignment
import pyshine as ps
import glob
from PyQt5 import QtCore, QtGui, QtWidgets
import exifread

#os.system("echo 123 | ")
#Untuk mengaktifkan i2c sendor jarak TOF
os.system("sudo chmod a+rw /dev/i2c-*")
warnings.filterwarnings("ignore") #digunakan untuk disabling warning message

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(2054, 1016)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.groupBox_3 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_3.setGeometry(QtCore.QRect(20, 0, 1801, 1011))
        self.groupBox_3.setObjectName("groupBox_3")
        self.groupBox_2 = QtWidgets.QGroupBox(self.groupBox_3)
        self.groupBox_2.setGeometry(QtCore.QRect(10, 20, 381, 71))
        self.groupBox_2.setObjectName("groupBox_2")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.groupBox_2)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.camList = QtWidgets.QComboBox(self.groupBox_2)
        self.camList.setMaximumSize(QtCore.QSize(350, 16777215))
        self.camList.setObjectName("camList")
        self.horizontalLayout.addWidget(self.camList)
        self.groupBox_4 = QtWidgets.QGroupBox(self.groupBox_3)
        self.groupBox_4.setGeometry(QtCore.QRect(10, 90, 781, 451))
        self.groupBox_4.setObjectName("groupBox_4")
        self.disp_main_1 = QtWidgets.QLabel(self.groupBox_4)
        self.disp_main_1.setGeometry(QtCore.QRect(30, 30, 721, 411))
        self.disp_main_1.setStyleSheet("background-color: rgb(246, 247, 247);")
        self.disp_main_1.setAlignment(QtCore.Qt.AlignCenter)
        self.disp_main_1.setObjectName("disp_main_1")
        self.label_2 = QtWidgets.QLabel(self.groupBox_3)
        self.label_2.setGeometry(QtCore.QRect(1570, 20, 81, 81))
        self.label_2.setText("")
        self.label_2.setPixmap(QtGui.QPixmap("/home/miconos/Documents/PMLD/LeafAreaMeter/Lambang UGM-hitam.png"))
        self.label_2.setScaledContents(True)
        self.label_2.setWordWrap(False)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.groupBox_3)
        self.label_3.setGeometry(QtCore.QRect(1660, 30, 121, 61))
        self.label_3.setText("")
        self.label_3.setPixmap(QtGui.QPixmap("/home/miconos/Documents/PMLD/LeafAreaMeter/logo-miconos.png"))
        self.label_3.setScaledContents(True)
        self.label_3.setObjectName("label_3")
        self.groupBox_9 = QtWidgets.QGroupBox(self.groupBox_3)
        self.groupBox_9.setGeometry(QtCore.QRect(10, 540, 1771, 261))
        self.groupBox_9.setObjectName("groupBox_9")
        self.disp_main_2 = QtWidgets.QLabel(self.groupBox_9)
        self.disp_main_2.setGeometry(QtCore.QRect(30, 70, 321, 181))
        self.disp_main_2.setMaximumSize(QtCore.QSize(511, 300))
        self.disp_main_2.setStyleSheet("background-color: rgb(246, 247, 247);")
        self.disp_main_2.setAlignment(QtCore.Qt.AlignCenter)
        self.disp_main_2.setObjectName("disp_main_2")
        self.disp_main_3 = QtWidgets.QLabel(self.groupBox_9)
        self.disp_main_3.setGeometry(QtCore.QRect(430, 70, 321, 181))
        self.disp_main_3.setMaximumSize(QtCore.QSize(511, 300))
        self.disp_main_3.setStyleSheet("background-color: rgb(246, 247, 247);")
        self.disp_main_3.setAlignment(QtCore.Qt.AlignCenter)
        self.disp_main_3.setObjectName("disp_main_3")
        self.disp_main_4 = QtWidgets.QLabel(self.groupBox_9)
        self.disp_main_4.setGeometry(QtCore.QRect(840, 70, 321, 181))
        self.disp_main_4.setMaximumSize(QtCore.QSize(511, 300))
        self.disp_main_4.setStyleSheet("background-color: rgb(246, 247, 247);")
        self.disp_main_4.setAlignment(QtCore.Qt.AlignCenter)
        self.disp_main_4.setObjectName("disp_main_4")
        self.disp_main_5 = QtWidgets.QLabel(self.groupBox_9)
        self.disp_main_5.setGeometry(QtCore.QRect(1260, 70, 321, 181))
        self.disp_main_5.setMaximumSize(QtCore.QSize(511, 300))
        self.disp_main_5.setStyleSheet("background-color: rgb(246, 247, 247);")
        self.disp_main_5.setAlignment(QtCore.Qt.AlignCenter)
        self.disp_main_5.setObjectName("disp_main_5")
        self.label = QtWidgets.QLabel(self.groupBox_9)
        self.label.setGeometry(QtCore.QRect(190, 30, 47, 13))
        self.label.setObjectName("label")
        self.label_10 = QtWidgets.QLabel(self.groupBox_9)
        self.label_10.setGeometry(QtCore.QRect(580, 30, 61, 16))
        self.label_10.setObjectName("label_10")
        self.label_12 = QtWidgets.QLabel(self.groupBox_9)
        self.label_12.setGeometry(QtCore.QRect(980, 40, 61, 16))
        self.label_12.setObjectName("label_12")
        self.label_13 = QtWidgets.QLabel(self.groupBox_9)
        self.label_13.setGeometry(QtCore.QRect(1330, 40, 191, 20))
        self.label_13.setObjectName("label_13")
        self.ckb_data_2 = QtWidgets.QCheckBox(self.groupBox_9)
        self.ckb_data_2.setGeometry(QtCore.QRect(1660, 60, 101, 31))
        self.ckb_data_2.setCheckable(True)
        self.ckb_data_2.setObjectName("ckb_data_2")
        self.btn_apply = QtWidgets.QPushButton(self.groupBox_9)
        self.btn_apply.setGeometry(QtCore.QRect(1660, 100, 100, 25))
        self.btn_apply.setMaximumSize(QtCore.QSize(100, 16777215))
        self.btn_apply.setObjectName("btn_apply")
        self.btn_restore = QtWidgets.QPushButton(self.groupBox_9)
        self.btn_restore.setGeometry(QtCore.QRect(1660, 140, 100, 25))
        self.btn_restore.setMaximumSize(QtCore.QSize(100, 16777215))
        self.btn_restore.setObjectName("btn_restore")
        self.display_clock = QtWidgets.QLCDNumber(self.groupBox_3)
        self.display_clock.setGeometry(QtCore.QRect(1400, 50, 151, 41))
        font = QtGui.QFont()
        font.setFamily("was@peenang")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_clock.setFont(font)
        self.display_clock.setMouseTracking(False)
        self.display_clock.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.display_clock.setAutoFillBackground(False)
        self.display_clock.setStyleSheet("color: rgb(255, 85, 0);")
        self.display_clock.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.display_clock.setFrameShadow(QtWidgets.QFrame.Raised)
        self.display_clock.setSmallDecimalPoint(True)
        self.display_clock.setDigitCount(8)
        self.display_clock.setObjectName("display_clock")
        self.groupBox_13 = QtWidgets.QGroupBox(self.groupBox_3)
        self.groupBox_13.setGeometry(QtCore.QRect(10, 820, 1771, 171))
        self.groupBox_13.setObjectName("groupBox_13")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout(self.groupBox_13)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.groupBox_6 = QtWidgets.QGroupBox(self.groupBox_13)
        self.groupBox_6.setMinimumSize(QtCore.QSize(228, 100))
        self.groupBox_6.setAutoFillBackground(False)
        self.groupBox_6.setFlat(False)
        self.groupBox_6.setCheckable(False)
        self.groupBox_6.setObjectName("groupBox_6")
        self.gridLayout_10 = QtWidgets.QGridLayout(self.groupBox_6)
        self.gridLayout_10.setObjectName("gridLayout_10")
        self.label_4 = QtWidgets.QLabel(self.groupBox_6)
        self.label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.label_4.setObjectName("label_4")
        self.gridLayout_10.addWidget(self.label_4, 0, 0, 1, 1)
        self.label_6 = QtWidgets.QLabel(self.groupBox_6)
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.label_6.setObjectName("label_6")
        self.gridLayout_10.addWidget(self.label_6, 0, 1, 1, 1)
        self.gridLayout_11 = QtWidgets.QGridLayout()
        self.gridLayout_11.setObjectName("gridLayout_11")
        self.label_7 = QtWidgets.QLabel(self.groupBox_6)
        self.label_7.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_7.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_7.setObjectName("label_7")
        self.gridLayout_11.addWidget(self.label_7, 0, 4, 1, 1)
        self.label_SMax_G = QtWidgets.QLabel(self.groupBox_6)
        self.label_SMax_G.setAlignment(QtCore.Qt.AlignCenter)
        self.label_SMax_G.setObjectName("label_SMax_G")
        self.gridLayout_11.addWidget(self.label_SMax_G, 1, 2, 1, 1)
        self.label_VMax_G = QtWidgets.QLabel(self.groupBox_6)
        self.label_VMax_G.setAlignment(QtCore.Qt.AlignCenter)
        self.label_VMax_G.setObjectName("label_VMax_G")
        self.gridLayout_11.addWidget(self.label_VMax_G, 2, 2, 1, 1)
        self.Slider_V_max_G = QtWidgets.QSlider(self.groupBox_6)
        self.Slider_V_max_G.setMinimumSize(QtCore.QSize(150, 18))
        self.Slider_V_max_G.setMaximumSize(QtCore.QSize(150, 18))
        self.Slider_V_max_G.setMaximum(255)
        self.Slider_V_max_G.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_V_max_G.setObjectName("Slider_V_max_G")
        self.gridLayout_11.addWidget(self.Slider_V_max_G, 2, 1, 1, 1)
        self.label_HMax_G = QtWidgets.QLabel(self.groupBox_6)
        self.label_HMax_G.setMinimumSize(QtCore.QSize(50, 0))
        self.label_HMax_G.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_HMax_G.setAlignment(QtCore.Qt.AlignCenter)
        self.label_HMax_G.setObjectName("label_HMax_G")
        self.gridLayout_11.addWidget(self.label_HMax_G, 0, 2, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(5, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_11.addItem(spacerItem, 0, 3, 1, 1)
        self.label_5 = QtWidgets.QLabel(self.groupBox_6)
        self.label_5.setMinimumSize(QtCore.QSize(10, 0))
        self.label_5.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_5.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_5.setObjectName("label_5")
        self.gridLayout_11.addWidget(self.label_5, 0, 0, 1, 1)
        self.label_8 = QtWidgets.QLabel(self.groupBox_6)
        self.label_8.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_8.setObjectName("label_8")
        self.gridLayout_11.addWidget(self.label_8, 1, 0, 1, 1)
        self.label_9 = QtWidgets.QLabel(self.groupBox_6)
        self.label_9.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_9.setObjectName("label_9")
        self.gridLayout_11.addWidget(self.label_9, 2, 0, 1, 1)
        self.label_11 = QtWidgets.QLabel(self.groupBox_6)
        self.label_11.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_11.setObjectName("label_11")
        self.gridLayout_11.addWidget(self.label_11, 1, 4, 1, 1)
        self.label_SMin_G = QtWidgets.QLabel(self.groupBox_6)
        self.label_SMin_G.setAlignment(QtCore.Qt.AlignCenter)
        self.label_SMin_G.setObjectName("label_SMin_G")
        self.gridLayout_11.addWidget(self.label_SMin_G, 1, 6, 1, 1)
        self.Slider_V_min_G = QtWidgets.QSlider(self.groupBox_6)
        self.Slider_V_min_G.setMinimumSize(QtCore.QSize(150, 18))
        self.Slider_V_min_G.setMaximumSize(QtCore.QSize(150, 18))
        self.Slider_V_min_G.setMaximum(255)
        self.Slider_V_min_G.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_V_min_G.setObjectName("Slider_V_min_G")
        self.gridLayout_11.addWidget(self.Slider_V_min_G, 2, 5, 1, 1)
        self.label_14 = QtWidgets.QLabel(self.groupBox_6)
        self.label_14.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_14.setObjectName("label_14")
        self.gridLayout_11.addWidget(self.label_14, 2, 4, 1, 1)
        self.label_VMin_G = QtWidgets.QLabel(self.groupBox_6)
        self.label_VMin_G.setAlignment(QtCore.Qt.AlignCenter)
        self.label_VMin_G.setObjectName("label_VMin_G")
        self.gridLayout_11.addWidget(self.label_VMin_G, 2, 6, 1, 1)
        self.Slider_S_min_G = QtWidgets.QSlider(self.groupBox_6)
        self.Slider_S_min_G.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_S_min_G.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_S_min_G.setMaximum(255)
        self.Slider_S_min_G.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_S_min_G.setObjectName("Slider_S_min_G")
        self.gridLayout_11.addWidget(self.Slider_S_min_G, 1, 5, 1, 1)
        self.Slider_H_min_G = QtWidgets.QSlider(self.groupBox_6)
        self.Slider_H_min_G.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_H_min_G.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_H_min_G.setMaximum(255)
        self.Slider_H_min_G.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_H_min_G.setTickPosition(QtWidgets.QSlider.NoTicks)
        self.Slider_H_min_G.setObjectName("Slider_H_min_G")
        self.gridLayout_11.addWidget(self.Slider_H_min_G, 0, 5, 1, 1)
        self.label_HMin_G = QtWidgets.QLabel(self.groupBox_6)
        self.label_HMin_G.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_HMin_G.setAlignment(QtCore.Qt.AlignCenter)
        self.label_HMin_G.setObjectName("label_HMin_G")
        self.gridLayout_11.addWidget(self.label_HMin_G, 0, 6, 1, 1)
        self.Slider_S_max_G = QtWidgets.QSlider(self.groupBox_6)
        self.Slider_S_max_G.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_S_max_G.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_S_max_G.setMaximum(255)
        self.Slider_S_max_G.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_S_max_G.setObjectName("Slider_S_max_G")
        self.gridLayout_11.addWidget(self.Slider_S_max_G, 1, 1, 1, 1)
        self.Slider_H_max_G = QtWidgets.QSlider(self.groupBox_6)
        self.Slider_H_max_G.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_H_max_G.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_H_max_G.setMaximum(255)
        self.Slider_H_max_G.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_H_max_G.setObjectName("Slider_H_max_G")
        self.gridLayout_11.addWidget(self.Slider_H_max_G, 0, 1, 1, 1)
        self.gridLayout_10.addLayout(self.gridLayout_11, 1, 0, 1, 2)
        self.horizontalLayout_2.addWidget(self.groupBox_6)
        self.groupBox_7 = QtWidgets.QGroupBox(self.groupBox_13)
        self.groupBox_7.setMinimumSize(QtCore.QSize(228, 100))
        self.groupBox_7.setAutoFillBackground(False)
        self.groupBox_7.setFlat(False)
        self.groupBox_7.setCheckable(False)
        self.groupBox_7.setObjectName("groupBox_7")
        self.gridLayout_12 = QtWidgets.QGridLayout(self.groupBox_7)
        self.gridLayout_12.setObjectName("gridLayout_12")
        self.label_19 = QtWidgets.QLabel(self.groupBox_7)
        self.label_19.setAlignment(QtCore.Qt.AlignCenter)
        self.label_19.setObjectName("label_19")
        self.gridLayout_12.addWidget(self.label_19, 0, 1, 1, 1)
        self.gridLayout_13 = QtWidgets.QGridLayout()
        self.gridLayout_13.setObjectName("gridLayout_13")
        self.label_20 = QtWidgets.QLabel(self.groupBox_7)
        self.label_20.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_20.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_20.setObjectName("label_20")
        self.gridLayout_13.addWidget(self.label_20, 0, 4, 1, 1)
        self.label_SMax_Y = QtWidgets.QLabel(self.groupBox_7)
        self.label_SMax_Y.setAlignment(QtCore.Qt.AlignCenter)
        self.label_SMax_Y.setObjectName("label_SMax_Y")
        self.gridLayout_13.addWidget(self.label_SMax_Y, 1, 2, 1, 1)
        self.label_VMax_Y = QtWidgets.QLabel(self.groupBox_7)
        self.label_VMax_Y.setAlignment(QtCore.Qt.AlignCenter)
        self.label_VMax_Y.setObjectName("label_VMax_Y")
        self.gridLayout_13.addWidget(self.label_VMax_Y, 2, 2, 1, 1)
        self.Slider_V_max_Y = QtWidgets.QSlider(self.groupBox_7)
        self.Slider_V_max_Y.setMinimumSize(QtCore.QSize(150, 18))
        self.Slider_V_max_Y.setMaximumSize(QtCore.QSize(150, 18))
        self.Slider_V_max_Y.setMaximum(255)
        self.Slider_V_max_Y.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_V_max_Y.setObjectName("Slider_V_max_Y")
        self.gridLayout_13.addWidget(self.Slider_V_max_Y, 2, 1, 1, 1)
        self.label_HMax_Y = QtWidgets.QLabel(self.groupBox_7)
        self.label_HMax_Y.setMinimumSize(QtCore.QSize(50, 0))
        self.label_HMax_Y.setMaximumSize(QtCore.QSize(20, 16777215))
        self.label_HMax_Y.setAlignment(QtCore.Qt.AlignCenter)
        self.label_HMax_Y.setObjectName("label_HMax_Y")
        self.gridLayout_13.addWidget(self.label_HMax_Y, 0, 2, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(5, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_13.addItem(spacerItem1, 0, 3, 1, 1)
        self.label_24 = QtWidgets.QLabel(self.groupBox_7)
        self.label_24.setMinimumSize(QtCore.QSize(50, 0))
        self.label_24.setMaximumSize(QtCore.QSize(10, 16777215))
        self.label_24.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_24.setObjectName("label_24")
        self.gridLayout_13.addWidget(self.label_24, 0, 0, 1, 1)
        self.label_25 = QtWidgets.QLabel(self.groupBox_7)
        self.label_25.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_25.setObjectName("label_25")
        self.gridLayout_13.addWidget(self.label_25, 1, 0, 1, 1)
        self.label_26 = QtWidgets.QLabel(self.groupBox_7)
        self.label_26.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_26.setObjectName("label_26")
        self.gridLayout_13.addWidget(self.label_26, 2, 0, 1, 1)
        self.label_27 = QtWidgets.QLabel(self.groupBox_7)
        self.label_27.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_27.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_27.setObjectName("label_27")
        self.gridLayout_13.addWidget(self.label_27, 1, 4, 1, 1)
        self.label_SMin_Y = QtWidgets.QLabel(self.groupBox_7)
        self.label_SMin_Y.setAlignment(QtCore.Qt.AlignCenter)
        self.label_SMin_Y.setObjectName("label_SMin_Y")
        self.gridLayout_13.addWidget(self.label_SMin_Y, 1, 6, 1, 1)
        self.Slider_V_min_Y = QtWidgets.QSlider(self.groupBox_7)
        self.Slider_V_min_Y.setMinimumSize(QtCore.QSize(150, 18))
        self.Slider_V_min_Y.setMaximumSize(QtCore.QSize(150, 18))
        self.Slider_V_min_Y.setMaximum(255)
        self.Slider_V_min_Y.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_V_min_Y.setObjectName("Slider_V_min_Y")
        self.gridLayout_13.addWidget(self.Slider_V_min_Y, 2, 5, 1, 1)
        self.label_29 = QtWidgets.QLabel(self.groupBox_7)
        self.label_29.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_29.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_29.setObjectName("label_29")
        self.gridLayout_13.addWidget(self.label_29, 2, 4, 1, 1)
        self.label_VMin_Y = QtWidgets.QLabel(self.groupBox_7)
        self.label_VMin_Y.setAlignment(QtCore.Qt.AlignCenter)
        self.label_VMin_Y.setObjectName("label_VMin_Y")
        self.gridLayout_13.addWidget(self.label_VMin_Y, 2, 6, 1, 1)
        self.Slider_S_min_Y = QtWidgets.QSlider(self.groupBox_7)
        self.Slider_S_min_Y.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_S_min_Y.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_S_min_Y.setMaximum(255)
        self.Slider_S_min_Y.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_S_min_Y.setObjectName("Slider_S_min_Y")
        self.gridLayout_13.addWidget(self.Slider_S_min_Y, 1, 5, 1, 1)
        self.Slider_H_min_Y = QtWidgets.QSlider(self.groupBox_7)
        self.Slider_H_min_Y.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_H_min_Y.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_H_min_Y.setMaximum(255)
        self.Slider_H_min_Y.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_H_min_Y.setTickPosition(QtWidgets.QSlider.NoTicks)
        self.Slider_H_min_Y.setObjectName("Slider_H_min_Y")
        self.gridLayout_13.addWidget(self.Slider_H_min_Y, 0, 5, 1, 1)
        self.label_HMin_Y = QtWidgets.QLabel(self.groupBox_7)
        self.label_HMin_Y.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_HMin_Y.setAlignment(QtCore.Qt.AlignCenter)
        self.label_HMin_Y.setObjectName("label_HMin_Y")
        self.gridLayout_13.addWidget(self.label_HMin_Y, 0, 6, 1, 1)
        self.Slider_S_max_Y = QtWidgets.QSlider(self.groupBox_7)
        self.Slider_S_max_Y.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_S_max_Y.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_S_max_Y.setMaximum(255)
        self.Slider_S_max_Y.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_S_max_Y.setObjectName("Slider_S_max_Y")
        self.gridLayout_13.addWidget(self.Slider_S_max_Y, 1, 1, 1, 1)
        self.Slider_H_max_Y = QtWidgets.QSlider(self.groupBox_7)
        self.Slider_H_max_Y.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_H_max_Y.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_H_max_Y.setMaximum(255)
        self.Slider_H_max_Y.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_H_max_Y.setObjectName("Slider_H_max_Y")
        self.gridLayout_13.addWidget(self.Slider_H_max_Y, 0, 1, 1, 1)
        self.gridLayout_12.addLayout(self.gridLayout_13, 1, 0, 1, 2)
        self.label_18 = QtWidgets.QLabel(self.groupBox_7)
        self.label_18.setAlignment(QtCore.Qt.AlignCenter)
        self.label_18.setObjectName("label_18")
        self.gridLayout_12.addWidget(self.label_18, 0, 0, 1, 1)
        self.horizontalLayout_2.addWidget(self.groupBox_7)
        self.groupBox_8 = QtWidgets.QGroupBox(self.groupBox_13)
        self.groupBox_8.setMinimumSize(QtCore.QSize(228, 100))
        self.groupBox_8.setAutoFillBackground(False)
        self.groupBox_8.setFlat(False)
        self.groupBox_8.setCheckable(False)
        self.groupBox_8.setObjectName("groupBox_8")
        self.gridLayout_14 = QtWidgets.QGridLayout(self.groupBox_8)
        self.gridLayout_14.setObjectName("gridLayout_14")
        self.label_32 = QtWidgets.QLabel(self.groupBox_8)
        self.label_32.setAlignment(QtCore.Qt.AlignCenter)
        self.label_32.setObjectName("label_32")
        self.gridLayout_14.addWidget(self.label_32, 0, 0, 1, 1)
        self.label_33 = QtWidgets.QLabel(self.groupBox_8)
        self.label_33.setAlignment(QtCore.Qt.AlignCenter)
        self.label_33.setObjectName("label_33")
        self.gridLayout_14.addWidget(self.label_33, 0, 1, 1, 1)
        self.gridLayout_15 = QtWidgets.QGridLayout()
        self.gridLayout_15.setObjectName("gridLayout_15")
        self.label_34 = QtWidgets.QLabel(self.groupBox_8)
        self.label_34.setMinimumSize(QtCore.QSize(50, 0))
        self.label_34.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_34.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_34.setObjectName("label_34")
        self.gridLayout_15.addWidget(self.label_34, 0, 4, 1, 1)
        self.label_SMax_B = QtWidgets.QLabel(self.groupBox_8)
        self.label_SMax_B.setAlignment(QtCore.Qt.AlignCenter)
        self.label_SMax_B.setObjectName("label_SMax_B")
        self.gridLayout_15.addWidget(self.label_SMax_B, 1, 2, 1, 1)
        self.label_VMax_B = QtWidgets.QLabel(self.groupBox_8)
        self.label_VMax_B.setAlignment(QtCore.Qt.AlignCenter)
        self.label_VMax_B.setObjectName("label_VMax_B")
        self.gridLayout_15.addWidget(self.label_VMax_B, 2, 2, 1, 1)
        self.Slider_V_max_B = QtWidgets.QSlider(self.groupBox_8)
        self.Slider_V_max_B.setMinimumSize(QtCore.QSize(150, 18))
        self.Slider_V_max_B.setMaximumSize(QtCore.QSize(150, 18))
        self.Slider_V_max_B.setMaximum(255)
        self.Slider_V_max_B.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_V_max_B.setObjectName("Slider_V_max_B")
        self.gridLayout_15.addWidget(self.Slider_V_max_B, 2, 1, 1, 1)
        self.label_HMax_B = QtWidgets.QLabel(self.groupBox_8)
        self.label_HMax_B.setMinimumSize(QtCore.QSize(50, 0))
        self.label_HMax_B.setMaximumSize(QtCore.QSize(20, 16777215))
        self.label_HMax_B.setAlignment(QtCore.Qt.AlignCenter)
        self.label_HMax_B.setObjectName("label_HMax_B")
        self.gridLayout_15.addWidget(self.label_HMax_B, 0, 2, 1, 1)
        spacerItem2 = QtWidgets.QSpacerItem(5, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_15.addItem(spacerItem2, 0, 3, 1, 1)
        self.label_38 = QtWidgets.QLabel(self.groupBox_8)
        self.label_38.setMinimumSize(QtCore.QSize(50, 0))
        self.label_38.setMaximumSize(QtCore.QSize(10, 16777215))
        self.label_38.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_38.setObjectName("label_38")
        self.gridLayout_15.addWidget(self.label_38, 0, 0, 1, 1)
        self.label_39 = QtWidgets.QLabel(self.groupBox_8)
        self.label_39.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_39.setObjectName("label_39")
        self.gridLayout_15.addWidget(self.label_39, 1, 0, 1, 1)
        self.label_40 = QtWidgets.QLabel(self.groupBox_8)
        self.label_40.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_40.setObjectName("label_40")
        self.gridLayout_15.addWidget(self.label_40, 2, 0, 1, 1)
        self.label_41 = QtWidgets.QLabel(self.groupBox_8)
        self.label_41.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_41.setObjectName("label_41")
        self.gridLayout_15.addWidget(self.label_41, 1, 4, 1, 1)
        self.label_SMin_B = QtWidgets.QLabel(self.groupBox_8)
        self.label_SMin_B.setAlignment(QtCore.Qt.AlignCenter)
        self.label_SMin_B.setObjectName("label_SMin_B")
        self.gridLayout_15.addWidget(self.label_SMin_B, 1, 6, 1, 1)
        self.Slider_V_min_B = QtWidgets.QSlider(self.groupBox_8)
        self.Slider_V_min_B.setMinimumSize(QtCore.QSize(150, 18))
        self.Slider_V_min_B.setMaximumSize(QtCore.QSize(150, 18))
        self.Slider_V_min_B.setMaximum(255)
        self.Slider_V_min_B.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_V_min_B.setObjectName("Slider_V_min_B")
        self.gridLayout_15.addWidget(self.Slider_V_min_B, 2, 5, 1, 1)
        self.label_43 = QtWidgets.QLabel(self.groupBox_8)
        self.label_43.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_43.setObjectName("label_43")
        self.gridLayout_15.addWidget(self.label_43, 2, 4, 1, 1)
        self.label_VMin_B = QtWidgets.QLabel(self.groupBox_8)
        self.label_VMin_B.setAlignment(QtCore.Qt.AlignCenter)
        self.label_VMin_B.setObjectName("label_VMin_B")
        self.gridLayout_15.addWidget(self.label_VMin_B, 2, 6, 1, 1)
        self.Slider_S_min_B = QtWidgets.QSlider(self.groupBox_8)
        self.Slider_S_min_B.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_S_min_B.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_S_min_B.setMaximum(255)
        self.Slider_S_min_B.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_S_min_B.setObjectName("Slider_S_min_B")
        self.gridLayout_15.addWidget(self.Slider_S_min_B, 1, 5, 1, 1)
        self.Slider_H_min_B = QtWidgets.QSlider(self.groupBox_8)
        self.Slider_H_min_B.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_H_min_B.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_H_min_B.setMaximum(255)
        self.Slider_H_min_B.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_H_min_B.setTickPosition(QtWidgets.QSlider.NoTicks)
        self.Slider_H_min_B.setObjectName("Slider_H_min_B")
        self.gridLayout_15.addWidget(self.Slider_H_min_B, 0, 5, 1, 1)
        self.label_HMin_B = QtWidgets.QLabel(self.groupBox_8)

        self.label_HMin_B.setMinimumSize(QtCore.QSize(30, 0))
        self.label_HMin_B.setMaximumSize(QtCore.QSize(50, 16777215))
        self.label_HMin_B.setAlignment(QtCore.Qt.AlignCenter)
        self.label_HMin_B.setObjectName("label_HMin_B")
        self.gridLayout_15.addWidget(self.label_HMin_B, 0, 6, 1, 1)
        self.Slider_S_max_B = QtWidgets.QSlider(self.groupBox_8)
        self.Slider_S_max_B.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_S_max_B.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_S_max_B.setMaximum(255)
        self.Slider_S_max_B.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_S_max_B.setObjectName("Slider_S_max_B")
        self.gridLayout_15.addWidget(self.Slider_S_max_B, 1, 1, 1, 1)
        self.Slider_H_max_B = QtWidgets.QSlider(self.groupBox_8)
        self.Slider_H_max_B.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_H_max_B.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_H_max_B.setMaximum(255)
        self.Slider_H_max_B.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_H_max_B.setObjectName("Slider_H_max_B")
        self.gridLayout_15.addWidget(self.Slider_H_max_B, 0, 1, 1, 1)
        self.gridLayout_14.addLayout(self.gridLayout_15, 1, 0, 1, 2)
        self.horizontalLayout_2.addWidget(self.groupBox_8)
        self.groupBox_11 = QtWidgets.QGroupBox(self.groupBox_13)
        self.groupBox_11.setMinimumSize(QtCore.QSize(228, 100))
        self.groupBox_11.setAutoFillBackground(False)
        self.groupBox_11.setFlat(False)
        self.groupBox_11.setCheckable(False)
        self.groupBox_11.setObjectName("groupBox_11")
        self.gridLayout_16 = QtWidgets.QGridLayout(self.groupBox_11)
        self.gridLayout_16.setObjectName("gridLayout_16")
        self.gridLayout_17 = QtWidgets.QGridLayout()
        self.gridLayout_17.setObjectName("gridLayout_17")
        self.Slider_kontras = QtWidgets.QSlider(self.groupBox_11)
        self.Slider_kontras.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_kontras.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_kontras.setMaximum(255)
        self.Slider_kontras.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_kontras.setObjectName("Slider_kontras")
        self.gridLayout_17.addWidget(self.Slider_kontras, 0, 1, 1, 1)
        self.Slider_kecerahan = QtWidgets.QSlider(self.groupBox_11)
        self.Slider_kecerahan.setMinimumSize(QtCore.QSize(150, 19))
        self.Slider_kecerahan.setMaximumSize(QtCore.QSize(150, 19))
        self.Slider_kecerahan.setMaximum(255)
        self.Slider_kecerahan.setOrientation(QtCore.Qt.Horizontal)
        self.Slider_kecerahan.setObjectName("Slider_kecerahan")
        self.gridLayout_17.addWidget(self.Slider_kecerahan, 1, 1, 1, 1)
        self.label_kontras = QtWidgets.QLabel(self.groupBox_11)
        self.label_kontras.setMinimumSize(QtCore.QSize(50, 0))
        self.label_kontras.setMaximumSize(QtCore.QSize(20, 16777215))
        self.label_kontras.setAlignment(QtCore.Qt.AlignCenter)
        self.label_kontras.setObjectName("label_kontras")
        self.gridLayout_17.addWidget(self.label_kontras, 0, 2, 1, 1)
        self.label_51 = QtWidgets.QLabel(self.groupBox_11)
        self.label_51.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_51.setObjectName("label_51")
        self.gridLayout_17.addWidget(self.label_51, 1, 0, 1, 1)
        self.label_kecerahan = QtWidgets.QLabel(self.groupBox_11)
        self.label_kecerahan.setAlignment(QtCore.Qt.AlignCenter)
        self.label_kecerahan.setObjectName("label_kecerahan")
        self.gridLayout_17.addWidget(self.label_kecerahan, 1, 2, 1, 1)
        self.label_52 = QtWidgets.QLabel(self.groupBox_11)
        self.label_52.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_52.setObjectName("label_52")
        self.gridLayout_17.addWidget(self.label_52, 0, 0, 1, 1)
        self.gridLayout_16.addLayout(self.gridLayout_17, 1, 0, 1, 2)
        self.label_35 = QtWidgets.QLabel(self.groupBox_11)
        self.label_35.setAlignment(QtCore.Qt.AlignCenter)
        self.label_35.setObjectName("label_35")
        self.gridLayout_16.addWidget(self.label_35, 0, 0, 1, 2)
        self.horizontalLayout_2.addWidget(self.groupBox_11)
        self.groupBox_14 = QtWidgets.QGroupBox(self.groupBox_3)
        self.groupBox_14.setGeometry(QtCore.QRect(800, 90, 981, 351))
        self.groupBox_14.setObjectName("groupBox_14")
        self.daungroup1 = QtWidgets.QGroupBox(self.groupBox_14)
        self.daungroup1.setGeometry(QtCore.QRect(330, 50, 321, 291))
        self.daungroup1.setObjectName("daungroup1")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.daungroup1)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.label_86 = QtWidgets.QLabel(self.daungroup1)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_86.setFont(font)
        self.label_86.setObjectName("label_86")
        self.gridLayout_2.addWidget(self.label_86, 1, 0, 1, 1)
        self.display_luas1 = QtWidgets.QLabel(self.daungroup1)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_luas1.setFont(font)
        self.display_luas1.setAlignment(QtCore.Qt.AlignCenter)
        self.display_luas1.setObjectName("display_luas1")
        self.gridLayout_2.addWidget(self.display_luas1, 2, 1, 1, 1)
        self.display_panjang1 = QtWidgets.QLabel(self.daungroup1)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_panjang1.setFont(font)
        self.display_panjang1.setAlignment(QtCore.Qt.AlignCenter)
        self.display_panjang1.setObjectName("display_panjang1")
        self.gridLayout_2.addWidget(self.display_panjang1, 0, 1, 1, 1)
        self.label_85 = QtWidgets.QLabel(self.daungroup1)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_85.setFont(font)
        self.label_85.setObjectName("label_85")
        self.gridLayout_2.addWidget(self.label_85, 0, 0, 1, 1)
        self.label_87 = QtWidgets.QLabel(self.daungroup1)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_87.setFont(font)
        self.label_87.setObjectName("label_87")
        self.gridLayout_2.addWidget(self.label_87, 2, 0, 1, 1)
        self.display_lebar1 = QtWidgets.QLabel(self.daungroup1)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_lebar1.setFont(font)
        self.display_lebar1.setAlignment(QtCore.Qt.AlignCenter)
        self.display_lebar1.setObjectName("display_lebar1")
        self.gridLayout_2.addWidget(self.display_lebar1, 1, 1, 1, 1)
        self.daungroup2 = QtWidgets.QGroupBox(self.groupBox_14)
        self.daungroup2.setGeometry(QtCore.QRect(660, 50, 311, 291))
        self.daungroup2.setObjectName("daungroup2")
        self.gridLayout_21 = QtWidgets.QGridLayout(self.daungroup2)
        self.gridLayout_21.setObjectName("gridLayout_21")
        self.label_105 = QtWidgets.QLabel(self.daungroup2)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_105.setFont(font)
        self.label_105.setObjectName("label_105")
        self.gridLayout_21.addWidget(self.label_105, 2, 0, 1, 1)
        self.label_103 = QtWidgets.QLabel(self.daungroup2)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_103.setFont(font)
        self.label_103.setObjectName("label_103")
        self.gridLayout_21.addWidget(self.label_103, 0, 0, 1, 1)
        self.label_104 = QtWidgets.QLabel(self.daungroup2)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_104.setFont(font)
        self.label_104.setObjectName("label_104")
        self.gridLayout_21.addWidget(self.label_104, 1, 0, 1, 1)
        self.display_panjang2 = QtWidgets.QLabel(self.daungroup2)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_panjang2.setFont(font)
        self.display_panjang2.setAlignment(QtCore.Qt.AlignCenter)
        self.display_panjang2.setObjectName("display_panjang2")
        self.gridLayout_21.addWidget(self.display_panjang2, 0, 1, 1, 1)
        self.display_luas2 = QtWidgets.QLabel(self.daungroup2)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_luas2.setFont(font)
        self.display_luas2.setAlignment(QtCore.Qt.AlignCenter)
        self.display_luas2.setObjectName("display_luas2")
        self.gridLayout_21.addWidget(self.display_luas2, 2, 1, 1, 1)
        self.display_lebar2 = QtWidgets.QLabel(self.daungroup2)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_lebar2.setFont(font)
        self.display_lebar2.setAlignment(QtCore.Qt.AlignCenter)
        self.display_lebar2.setObjectName("display_lebar2")
        self.gridLayout_21.addWidget(self.display_lebar2, 1, 1, 1, 1)
        self.label_112 = QtWidgets.QLabel(self.groupBox_14)
        self.label_112.setGeometry(QtCore.QRect(10, 30, 67, 17))
        self.label_112.setObjectName("label_112")
        self.label_113 = QtWidgets.QLabel(self.groupBox_14)
        self.label_113.setGeometry(QtCore.QRect(330, 30, 67, 17))
        self.label_113.setObjectName("label_113")
        self.label_114 = QtWidgets.QLabel(self.groupBox_14)
        self.label_114.setGeometry(QtCore.QRect(660, 30, 67, 17))
        self.label_114.setObjectName("label_114")
        self.daungroup0 = QtWidgets.QGroupBox(self.groupBox_14)
        self.daungroup0.setGeometry(QtCore.QRect(10, 50, 311, 291))
        self.daungroup0.setObjectName("daungroup0")
        self.gridLayout = QtWidgets.QGridLayout(self.daungroup0)
        self.gridLayout.setObjectName("gridLayout")
        self.display_luas0 = QtWidgets.QLabel(self.daungroup0)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_luas0.setFont(font)
        self.display_luas0.setAlignment(QtCore.Qt.AlignCenter)
        self.display_luas0.setObjectName("display_luas0")
        self.gridLayout.addWidget(self.display_luas0, 2, 1, 1, 1)
        self.display_panjang0 = QtWidgets.QLabel(self.daungroup0)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_panjang0.setFont(font)
        self.display_panjang0.setAlignment(QtCore.Qt.AlignCenter)
        self.display_panjang0.setObjectName("display_panjang0")
        self.gridLayout.addWidget(self.display_panjang0, 0, 1, 1, 1)
        self.label_17 = QtWidgets.QLabel(self.daungroup0)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_17.setFont(font)
        self.label_17.setObjectName("label_17")
        self.gridLayout.addWidget(self.label_17, 1, 0, 1, 1)
        self.display_lebar0 = QtWidgets.QLabel(self.daungroup0)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.display_lebar0.setFont(font)
        self.display_lebar0.setAlignment(QtCore.Qt.AlignCenter)
        self.display_lebar0.setObjectName("display_lebar0")
        self.gridLayout.addWidget(self.display_lebar0, 1, 1, 1, 1)
        self.label_36 = QtWidgets.QLabel(self.daungroup0)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_36.setFont(font)
        self.label_36.setObjectName("label_36")
        self.gridLayout.addWidget(self.label_36, 2, 0, 1, 1)
        self.label_15 = QtWidgets.QLabel(self.daungroup0)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_15.setFont(font)
        self.label_15.setObjectName("label_15")
        self.gridLayout.addWidget(self.label_15, 0, 0, 1, 1)
        self.daungroup0.raise_()
        self.daungroup1.raise_()
        self.daungroup2.raise_()
        self.label_112.raise_()
        self.label_114.raise_()
        self.label_113.raise_() # lower
        self.groupBox_5 = QtWidgets.QGroupBox(self.groupBox_3)
        self.groupBox_5.setGeometry(QtCore.QRect(800, 440, 981, 101))
        self.groupBox_5.setObjectName("groupBox_5")
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout(self.groupBox_5)
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.textEdit = QtWidgets.QTextEdit(self.groupBox_5)
        self.textEdit.setStyleSheet("background-color: rgb(246, 247, 247);\n"
"border-radius:3px;")
        self.textEdit.setReadOnly(True)
        self.textEdit.setObjectName("textEdit")
        self.horizontalLayout_4.addWidget(self.textEdit)
        self.btn_save = QtWidgets.QPushButton(self.groupBox_5)
        self.btn_save.setMaximumSize(QtCore.QSize(100, 16777215))
        self.btn_save.setObjectName("btn_save")
        self.horizontalLayout_4.addWidget(self.btn_save)
        self.groupBox_12 = QtWidgets.QGroupBox(self.groupBox_3)
        self.groupBox_12.setGeometry(QtCore.QRect(1050, 20, 331, 71))
        self.groupBox_12.setObjectName("groupBox_12")
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout(self.groupBox_12)
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.gridLayout_8 = QtWidgets.QGridLayout()
        self.gridLayout_8.setSizeConstraint(QtWidgets.QLayout.SetNoConstraint)
        self.gridLayout_8.setHorizontalSpacing(6)
        self.gridLayout_8.setVerticalSpacing(0)
        self.gridLayout_8.setObjectName("gridLayout_8")
        self.label_16 = QtWidgets.QLabel(self.groupBox_12)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_16.setFont(font)
        self.label_16.setAlignment(QtCore.Qt.AlignCenter)
        self.label_16.setObjectName("label_16")
        self.gridLayout_8.addWidget(self.label_16, 0, 1, 1, 1)
        self.Qlabel_cpu = QtWidgets.QLabel(self.groupBox_12)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.Qlabel_cpu.setFont(font)
        self.Qlabel_cpu.setAlignment(QtCore.Qt.AlignCenter)
        self.Qlabel_cpu.setObjectName("Qlabel_cpu")
        self.gridLayout_8.addWidget(self.Qlabel_cpu, 1, 1, 1, 1)
        self.Qlabel_ram = QtWidgets.QLabel(self.groupBox_12)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.Qlabel_ram.setFont(font)
        self.Qlabel_ram.setAlignment(QtCore.Qt.AlignCenter)
        self.Qlabel_ram.setObjectName("Qlabel_ram")
        self.gridLayout_8.addWidget(self.Qlabel_ram, 1, 2, 1, 1)
        self.label_28 = QtWidgets.QLabel(self.groupBox_12)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_28.setFont(font)
        self.label_28.setAlignment(QtCore.Qt.AlignCenter)
        self.label_28.setObjectName("label_28")
        self.gridLayout_8.addWidget(self.label_28, 0, 2, 1, 1)
        self.Qlabel_fps = QtWidgets.QLabel(self.groupBox_12)
        font = QtGui.QFont()
        font.setFamily("Droid Sans Fallback")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.Qlabel_fps.setFont(font)
        self.Qlabel_fps.setStyleSheet("color: rgb(237, 85, 59);")
        self.Qlabel_fps.setAlignment(QtCore.Qt.AlignCenter)
        self.Qlabel_fps.setObjectName("Qlabel_fps")
        self.gridLayout_8.addWidget(self.Qlabel_fps, 1, 0, 1, 1)
        self.label_21 = QtWidgets.QLabel(self.groupBox_12)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_21.setFont(font)
        self.label_21.setAlignment(QtCore.Qt.AlignCenter)
        self.label_21.setObjectName("label_21")
        self.gridLayout_8.addWidget(self.label_21, 0, 0, 1, 1)
        self.horizontalLayout_5.addLayout(self.gridLayout_8)
        self.groupBox_10 = QtWidgets.QGroupBox(self.groupBox_3)
        self.groupBox_10.setGeometry(QtCore.QRect(410, 20, 621, 71))
        self.groupBox_10.setObjectName("groupBox_10")
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout(self.groupBox_10)
        self.horizontalLayout_6.setObjectName("horizontalLayout_6")
        self.btn_start_1 = QtWidgets.QPushButton(self.groupBox_10)
        self.btn_start_1.setMaximumSize(QtCore.QSize(130, 16777215))
        self.btn_start_1.setObjectName("btn_start_1")
        self.horizontalLayout_6.addWidget(self.btn_start_1)
        self.btn_start_2 = QtWidgets.QPushButton(self.groupBox_10)
        self.btn_start_2.setMaximumSize(QtCore.QSize(120, 16777215))
        self.btn_start_2.setObjectName("btn_start_2")
        self.horizontalLayout_6.addWidget(self.btn_start_2)
        self.btn_start_3 = QtWidgets.QPushButton(self.groupBox_10)
        self.btn_start_3.setMaximumSize(QtCore.QSize(90, 16777215))
        self.btn_start_3.setObjectName("btn_start_3")
        self.horizontalLayout_6.addWidget(self.btn_start_3)
        self.btn_not_ai = QtWidgets.QPushButton(self.groupBox_10)
        self.btn_not_ai.setEnabled(False)
        self.btn_not_ai.setMaximumSize(QtCore.QSize(90, 16777215))
        self.btn_not_ai.setObjectName("btn_not_ai")
        self.horizontalLayout_6.addWidget(self.btn_not_ai)
        self.btn_calibration = QtWidgets.QPushButton(self.groupBox_10)
        self.btn_calibration.setEnabled(False)
        self.btn_calibration.setMaximumSize(QtCore.QSize(100, 16777215))
        self.btn_calibration.setObjectName("btn_calibration")
        self.horizontalLayout_6.addWidget(self.btn_calibration)
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.DateTime = QDateTime.currentDateTime()
        self.display_clock.display(self.DateTime.toString('hh:mm:ss'))

        #Pilih Camera
        self.camList.addItem("C2CMOS02100KPA: C2CMOS0200KPA")
        self.online_cam = QCameraInfo.availableCameras()
        self.camList.addItems([c.description() for c in self.online_cam])
        camIndex = self.camList.currentIndex()
        self.textEdit.append(f"{self.DateTime.toString('d MMMM yy hh:mm:ss')}: Webcam ({self.camList.currentText()})")

        self.file_name = "data_daun.xlsx"
        self.tmp = None # Hold the temporary image for display
        self.started = False  
        self.filename = 'Snapshot '+str(time.strftime('%Y-%b-%d at %H.%M.%S %p'))+'.png'
        self.freeze=False
        self.btn_save.setEnabled(False)

        self.btn_not_ai.clicked.connect(self.tombol_not_use_ai)
        self.btn_start_1.clicked.connect(self.tombol_ai)      #supaya tombol start stop use / not use ai bisa berfungsi
        self.btn_start_2.clicked.connect(self.savePhoto) 
        self.btn_start_3.clicked.connect(self.ukurDaun)
        self.btn_apply.clicked.connect(self.klik_apply)  
        self.btn_restore.clicked.connect(self.klik_restore)
        self.btn_save.clicked.connect(self.save_daun)      
        self.ckb_data_2.stateChanged.connect(self.slider)

    #PILIHAN SEGMENTASI MANUAL
    def tombol_not_use_ai(self):
        if self.started:
            self.started=False 
            self.btn_not_ai.setText('START')
        else:
            self.started=True
            self.btn_not_ai.setText('STOP') 
            self.update()
        
        cam = cv2.VideoCapture(0) #camera isi 0
        while(cam.isOpened()):
            QtWidgets.QApplication.processEvents()
            img, self.image = cam.read()
            self.image = imutils.resize(self.image, height = 300)


            self.update()
            if self.started==False:
                break
  
    #PILIHAN SEGMENTASI MENGGUNAKAN AI
    def tombol_ai(self):
        #if self.ckb_ai.isChecked(): #jika checkbox ai di checklist
        if self.started:
            self.DateTime = QDateTime.currentDateTime()
            self.display_clock.display(self.DateTime.toString('hh:mm:ss'))
            self.started=False 
            self.btn_start_1.setText('START')
            self.freeze=True
            self.btn_save.setEnabled(True)

            #self.textEdit.setText('Camera Pause') 
            self.DateTime = QDateTime.currentDateTime()
            self.textEdit.append(f"{self.DateTime.toString('d MMMM yy hh:mm:ss')}: Camera Pause")
        else:
            self.DateTime = QDateTime.currentDateTime()
            self.display_clock.display(self.DateTime.toString('hh:mm:ss'))
            self.started=True
            self.btn_start_1.setText('PAUSE') 
            self.freeze=False
            self.btn_save.setEnabled(False)

            self.DateTime = QDateTime.currentDateTime()
            self.textEdit.append(f"{self.DateTime.toString('d MMMM yy hh:mm:ss')}: Camera ON")
        
        cam = cv2.VideoCapture(0) #camera isi 0
        while(cam.isOpened()):
            QtWidgets.QApplication.processEvents()
            img, self.image = cam.read()
            self.image = imutils.resize(self.image, height = 300)
            self.update() 
            if self.started == False:
                self.ukurJarak()
                break
                print('Loop break')

    #def pilih_kamera(self):

    def normalisasi_cahaya(self):
        #illumination normalize
        ycrcb = cv2.cvtColor(self.image, cv2.COLOR_BGR2YCrCb)
        y, cr, cb = cv2.split(ycrcb)

        gaussian = cv2.GaussianBlur(y, (0, 0), 5,5)
        y = y - gaussian
        ycrcb = cv2.merge([y, cr, cb])

        self.hasil_normalisasi = cv2.cvtColor(ycrcb, cv2.COLOR_YCrCb2BGR)

    def ukurJarak(self):
        #pengukuran menggunakan sensor jarak
        i2c =busio.I2C(board.SCL, board.SDA)
        vl53 = adafruit_vl53l0x.VL53L0X(i2c)
        jarak = float(format(vl53.range))
        jarak2 = jarak + 150.0
        print(jarak2)
        time.sleep(1)
        self.jarak2 = jarak2

    def savePhoto(self):
        self.DateTime = QDateTime.currentDateTime()
        self.display_clock.display(self.DateTime.toString('hh:mm:ss'))
        #menyimpan hasil ambil gambar
        folderSnapshot='/home/miconos/Documents/PMLD/LeafAreaMeter/Snapshot/'
        self.outname = str(time.strftime('%Y-%b-%d at %H.%M.%S %p'))+'.png'
        self.filename = folderSnapshot + 'Snapshot' + self.outname
        cv2.imwrite(self.filename,self.tmp)
        print('Image saved as: ', self.filename)
        self.DateTime = QDateTime.currentDateTime()
        self.textEdit.append(f"{self.DateTime.toString('d MMMM yy hh:mm:ss')}: Image Captured")

    def ukurDaun(self):
        # Display current time on the UI
        self.DateTime = QDateTime.currentDateTime()
        self.display_clock.display(self.DateTime.toString('hh:mm:ss'))

        # Normalize lighting (assuming this function is defined elsewhere in your code)
        self.normalisasi_cahaya()

        # Define file paths
        folder_path = '/home/miconos/Documents/PMLD/LeafAreaMeter/Snapshot/'  # Source image
        folder_mask = '/home/miconos/Documents/PMLD/LeafAreaMeter/Masks/'
        folder_outputs = '/home/miconos/Documents/PMLD/LeafAreaMeter/Outputs/'
        folder_daun = '/home/miconos/Documents/PMLD/LeafAreaMeter/Daun/'
        folder_runs = '/home/miconos/Documents/PMLD/LeafAreaMeter/runs/'  # Adjusted to match consistency
        res_name = self.outname  # Output name for saving results

        # Locate the latest snapshot
        list_of_files = glob.glob(os.path.join(folder_path, '*.png'))
        latest_file = max(list_of_files, key=os.path.getctime)
        image = cv2.imread(latest_file)
        self.image2 = image

        # Check image resolution
        resolusi = image.shape

        # Load YOLO model
        model = YOLO('/home/miconos/Documents/PMLD/LeafAreaMeter/daun200.pt')

        # Predict with YOLO
        results = model.predict(
            self.image,
            conf=0.55,
            save=True,
            show_boxes=True,
            show_labels=False
        )

        # Prepare original image and combined mask
        self.image2 = cv2.imread(latest_file)
        combined_mask = np.zeros((self.image2.shape[0], self.image2.shape[1]), dtype=np.uint8)

        # Initialize lists to store measurements for up to 3 leaves
        self.lebarobjek = []
        self.panjangobjek = []
        self.luasobjek = [] 
        self.piksel_w_objek = []
        self.piksel_h_objek = [] 
        self.luas_piksel = [] 

        # Process each bounding box, limiting to 3 total leaves
        leaf_detected = False
        self.objectdetected = 0
        for result in results[:3]:  # Limit to 3 results
            boxes = result.boxes
            masks = result.masks

            if boxes:
                leaf_detected = True
                self.objectdetected += len(boxes)
                for i, box in enumerate(boxes):
                    # Get bounding box dimensions (center x, y, width, height)
                    x, y, w, h = (float(coord) for coord in box.xywh[0])
                    
                    # Dimensions of image in pixels
                    lebar_citra = 400
                    panjang_citra = 300

                    # Sensor distance and angle values
                    dsensor = self.jarak2
                    theta = 16
                    alpha = 12

                    # Real-world dimensions
                    sudut = math.tan(math.radians(theta / 2))
                    sudut2 = math.tan(math.radians(alpha / 2))

                    nilaiH = 2 * dsensor * sudut
                    nilaiW = 2 * dsensor * sudut2

                    # Calculate real dimensions of object
                    lebarobjek_value = abs(round((nilaiW * w) / lebar_citra, 3))
                    panjangobjek_value = abs(round((nilaiH * h) / panjang_citra, 3))
                    luasobjek_value = abs(round((lebarobjek_value * panjangobjek_value), 3))
                    piksel_w_objek_value = (round(w))
                    piksel_h_objek_value = (round(h))
                    luas_piksel_value = abs(round((w * h), 3))

                    # Append measurements to lists
                    self.lebarobjek.append(lebarobjek_value)
                    self.panjangobjek.append(panjangobjek_value)
                    self.luasobjek.append(luasobjek_value)
                    self.piksel_w_objek.append(piksel_w_objek_value)
                    self.piksel_h_objek.append(piksel_h_objek_value)
                    self.luas_piksel.append(luas_piksel_value)

                    # Output results to terminal
                    print(f"Bounding Box: {x:.0f} px x {y:.0f} px")
                    print(f"Width: {lebarobjek_value} mm, Height: {panjangobjek_value} mm, Area: {luasobjek_value} mm²")
                    print(f"Pixel width: {piksel_w_objek_value}\n")
                    print(f"Pixel high: {piksel_h_objek_value}\n")
                    print(f"luas piksel {luas_piksel_value}\n")

                    # Add mask if available
                    if masks is not None:
                        mask_data = masks.data[i].numpy()
                        mask_data = (mask_data * 255).astype(np.uint8)
                        mask_resized = cv2.resize(mask_data, (self.image2.shape[1], self.image2.shape[0]))
                        combined_mask = cv2.bitwise_or(combined_mask, mask_resized)

        # Check if any leaf was detected and log accordingly
        if not leaf_detected:
            self.DateTime = QDateTime.currentDateTime()
            self.textEdit.append(f"{self.DateTime.toString('d MMMM yy hh:mm:ss')}: Leaf Not Detected")
            print("Tidak Ada Daun")
        else:
            # Save combined mask in black-and-white format
            mask = (combined_mask > 0).astype(np.uint8) * 255
            self.daunmask = mask
            cv2.imwrite(os.path.join(folder_mask, f'mask_{res_name}'), mask)

            # Apply combined mask to original image
            self.masked = cv2.bitwise_and(self.image2, self.image2, mask=combined_mask)
            cv2.imwrite(os.path.join(folder_daun, f'daun_{res_name}'), self.masked)

            # Copy predicted image to outputs
            list_of_file2 = glob.glob(os.path.join(folder_outputs, "*.png"))
            self.latest_file2 = max(list_of_file2, key=os.path.getctime)
            print(self.latest_file2)

        # Check if the predicted image is loaded correctly
        predicted_image_path = os.path.join(folder_runs, "segment", "predict", "image0.jpg")
        predicted_image = cv2.imread(predicted_image_path)

        # Verify that the image is loaded
        if predicted_image is None:
            print(f"Error: Predicted image {predicted_image_path} not loaded properly.")
        else:
            # Annotate the predicted image with object count
            for result in results:
                boxes = result.boxes
                for self.objectdetected, box in enumerate(boxes, start=1):
                    x, y, w, h = (float(coord) for coord in box.xywh[0])

                    # Convert to pixel coordinates for labeling
                    center_x, center_y = int(x), int(y)

                    # Add label to the predicted image
                    cv2.putText(
                        predicted_image,
                        f"{self.objectdetected}",
                        (center_x-10, center_y),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        1.3,  # Font scale
                        (255, 255, 255),  # Font color (white)
                        3,  # Thickness
                        cv2.LINE_AA
                    )

            # Save annotated image
            annotated_path = os.path.join(folder_outputs, f"predict_{res_name}")
            cv2.imwrite(annotated_path, predicted_image)

            # Resize the image using imutils
            resized_image = imutils.resize(predicted_image, height=411)  # Adjust height as needed

            # Convert resized image to RGB (Qt requires RGB format)
            resized_image_rgb = cv2.cvtColor(resized_image, cv2.COLOR_BGR2RGB)

            # Convert to QImage
            height, width, channel = resized_image_rgb.shape
            bytes_per_line = channel * width
            q_image = QImage(
                resized_image_rgb.data,
                width,
                height,
                bytes_per_line,
                QImage.Format_RGB888
            )

            # Set the QImage as pixmap for the display widget
            pixmap = QPixmap.fromImage(q_image)
            self.disp_main_1.setPixmap(pixmap)
            print("Resized image displayed successfully.")

            
            shutil.rmtree(folder_runs, ignore_errors=True)

            # Log successful analysis
            self.DateTime = QDateTime.currentDateTime()
            self.textEdit.append(f"{self.DateTime.toString('d MMMM yy hh:mm:ss')}: Successful Leaf Analysis")

        # Update UI with new data
        self.ckb_data_2.setChecked(False)
        # self.update()
        self.displayData()
        # self.disp_main_1.setPixmap(QPixmap.fromImage())


    #menampilkan streaming video ke GUI
    def tampilkankamera(self,image):
        #display di label
        self.tmp = image
        image = imutils.resize(image,width=720) #ukurannya pas GUI
        frame = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        image = QImage(frame, frame.shape[1],frame.shape[0],frame.strides[0],QImage.Format_RGB888)
        self.disp_main_1.setPixmap(QtGui.QPixmap.fromImage(image))
             
    def update(self):
        img = self.image  
        self.tampilkankamera(img)
    
    def update_hsv(self, lower_green, upper_green, lower_yellow, upper_yellow, lower_brown, upper_brown, contrast, brightness):

        self.lower_green = lower_green
        self.upper_green = upper_green
        self.lower_yellow = lower_yellow
        self.upper_yellow = upper_yellow
        self.lower_brown = lower_brown
        self.upper_brown = upper_brown
        self.contrast = contrast
        self.brightness = brightness

    #read file txt slider
    def readFromFile(self, value):
        try:
            with open('slider_value.txt', 'r') as file:
                lines = file.readline().split(",")
                return int(lines[value])
                print()
        except (FileNotFoundError, ValueError):
            return 50   

    #Untuk slider. Baris : 508
    def slider(self):
        if self.ckb_data_2.isChecked():
            self.DateTime = QDateTime.currentDateTime()
            self.display_clock.display(self.DateTime.toString('hh:mm:ss'))
            # Parameter Green
            self.Slider_H_max_G.setValue(self.readFromFile(0))
            self.Slider_S_max_G.setValue(self.readFromFile(1))
            self.Slider_V_max_G.setValue(self.readFromFile(2))
            self.Slider_H_min_G.setValue(self.readFromFile(3))
            self.Slider_S_min_G.setValue(self.readFromFile(4))
            self.Slider_V_min_G.setValue(self.readFromFile(5))
            
            # Parameter Yellow
            self.Slider_H_max_Y.setValue(self.readFromFile(6))
            self.Slider_S_max_Y.setValue(self.readFromFile(7))
            self.Slider_V_max_Y.setValue(self.readFromFile(8))
            self.Slider_H_min_Y.setValue(self.readFromFile(9))
            self.Slider_S_min_Y.setValue(self.readFromFile(10))
            self.Slider_V_min_Y.setValue(self.readFromFile(11))

            #parameter Brown
            self.Slider_H_max_B.setValue(self.readFromFile(12))
            self.Slider_S_max_B.setValue(self.readFromFile(13))
            self.Slider_V_max_B.setValue(self.readFromFile(14))
            self.Slider_H_min_B.setValue(self.readFromFile(15))
            self.Slider_S_min_B.setValue(self.readFromFile(16))
            self.Slider_V_min_B.setValue(self.readFromFile(17))

            self.Slider_kontras.setValue(self.readFromFile(18))
            self.Slider_kecerahan.setValue(self.readFromFile(19))

            print("file index 0 : ", self.readFromFile(0))
            print("file index 1 : ", self.readFromFile(1))
            print("file index 2 : ", self.readFromFile(2))
            print("file index 3 : ", self.readFromFile(3))
            print("file index 4 : ", self.readFromFile(5))
            print("file index 5 : ", self.readFromFile(6))
            print("file index 6 : ", self.readFromFile(7))
            print("file index 7 : ", self.readFromFile(8))
            print("file index 8 : ", self.readFromFile(9))
            print("file index 9 : ", self.readFromFile(10))
            print("file index 10 : ", self.readFromFile(11))
            print("file index 11 : ", self.readFromFile(12))
            print("file index 12 : ", self.readFromFile(13))
            print("file index 13 : ", self.readFromFile(14))
            print("file index 14 : ", self.readFromFile(15))
            print("file index 15 : ", self.readFromFile(16))
            print("file index 16 : ", self.readFromFile(17))
            print("file index 17 : ", self.readFromFile(18))
            print("file index 18 : ", self.readFromFile(19))
            print("file index 19 : ", self.readFromFile(20))

            #Parameter Hijau
            self.Slider_H_max_G.valueChanged.connect(self.sliderChanged)
            self.Slider_S_max_G.valueChanged.connect(self.sliderChanged)
            self.Slider_V_max_G.valueChanged.connect(self.sliderChanged)
            self.Slider_H_min_G.valueChanged.connect(self.sliderChanged)
            self.Slider_S_min_G.valueChanged.connect(self.sliderChanged)
            self.Slider_V_min_G.valueChanged.connect(self.sliderChanged)

            # Parameter Kuning
            self.Slider_H_max_Y.valueChanged.connect(self.sliderChanged)
            self.Slider_S_max_Y.valueChanged.connect(self.sliderChanged)
            self.Slider_V_max_Y.valueChanged.connect(self.sliderChanged)
            self.Slider_H_min_Y.valueChanged.connect(self.sliderChanged)
            self.Slider_S_min_Y.valueChanged.connect(self.sliderChanged)
            self.Slider_V_min_Y.valueChanged.connect(self.sliderChanged)

            # Parameter Coklat
            self.Slider_H_max_B.valueChanged.connect(self.sliderChanged)
            self.Slider_S_max_B.valueChanged.connect(self.sliderChanged)
            self.Slider_V_max_B.valueChanged.connect(self.sliderChanged)
            self.Slider_H_min_B.valueChanged.connect(self.sliderChanged)
            self.Slider_S_min_B.valueChanged.connect(self.sliderChanged)
            self.Slider_V_min_B.valueChanged.connect(self.sliderChanged)

            self.Slider_kontras.valueChanged.connect(self.sliderChanged)
            self.Slider_kecerahan.valueChanged.connect(self.sliderChanged)

            self.lower_green = np.array([self.Slider_H_min_G.value(
            ), self.Slider_S_min_G.value(), self.Slider_V_min_G.value()])
            self.upper_green = np.array([self.Slider_H_max_G.value(
            ), self.Slider_S_max_G.value(), self.Slider_V_max_G.value()])

            self.lower_yellow = np.array([self.Slider_H_min_Y.value(
            ), self.Slider_S_min_Y.value(), self.Slider_V_min_Y.value()])
            self.upper_yellow = np.array([self.Slider_H_max_Y.value(
            ), self.Slider_S_max_Y.value(), self.Slider_V_max_Y.value()])

            self.lower_brown = np.array([self.Slider_H_min_B.value(
            ), self.Slider_S_min_B.value(), self.Slider_V_min_B.value()])
            self.upper_brown = np.array([self.Slider_H_max_B.value(
            ), self.Slider_S_max_B.value(), self.Slider_V_max_B.value()])

            #Menampilkan angka slider di GUI
            self.label_HMax_G.setText(f"{self.Slider_H_max_G.value()}")
            self.label_SMax_G.setText(f"{self.Slider_S_max_G.value()}")
            self.label_VMax_G.setText(f"{self.Slider_V_max_G.value()}")
            self.label_HMin_G.setText(f"{self.Slider_H_min_G.value()}")
            self.label_SMin_G.setText(f"{self.Slider_S_min_G.value()}")
            self.label_VMin_G.setText(f"{self.Slider_V_min_G.value()}")

            self.label_HMax_Y.setText(f"{self.Slider_H_max_Y.value()}")
            self.label_SMax_Y.setText(f"{self.Slider_S_max_Y.value()}")
            self.label_VMax_Y.setText(f"{self.Slider_V_max_Y.value()}")
            self.label_HMin_Y.setText(f"{self.Slider_H_min_Y.value()}")
            self.label_SMin_Y.setText(f"{self.Slider_S_min_Y.value()}")
            self.label_VMin_Y.setText(f"{self.Slider_V_min_Y.value()}")

            self.label_HMax_B.setText(f"{self.Slider_H_max_B.value()}")
            self.label_SMax_B.setText(f"{self.Slider_S_max_B.value()}")
            self.label_VMax_B.setText(f"{self.Slider_V_max_B.value()}")
            self.label_HMin_B.setText(f"{self.Slider_H_min_B.value()}")
            self.label_SMin_B.setText(f"{self.Slider_S_min_B.value()}")
            self.label_VMin_B.setText(f"{self.Slider_V_min_B.value()}")

            self.label_kontras.setText(f"{self.Slider_kontras.value()}")
            self.label_kecerahan.setText(f"{self.Slider_kecerahan.value()}")
            self.displayManualAdjustmentResult()

    def sliderChanged(self):
        self.on_value_change()
        self.saveToFile()
        self.label_HMax_G.setText(f"{self.Slider_H_max_G.value()}")
        self.label_SMax_G.setText(f"{self.Slider_S_max_G.value()}")
        self.label_VMax_G.setText(f"{self.Slider_V_max_G.value()}")
        self.label_HMin_G.setText(f"{self.Slider_H_min_G.value()}")
        self.label_SMin_G.setText(f"{self.Slider_S_min_G.value()}")
        self.label_VMin_G.setText(f"{self.Slider_V_min_G.value()}")

        self.label_HMax_Y.setText(f"{self.Slider_H_max_Y.value()}")
        self.label_SMax_Y.setText(f"{self.Slider_S_max_Y.value()}")
        self.label_VMax_Y.setText(f"{self.Slider_V_max_Y.value()}")
        self.label_HMin_Y.setText(f"{self.Slider_H_min_Y.value()}")
        self.label_SMin_Y.setText(f"{self.Slider_S_min_Y.value()}")
        self.label_VMin_Y.setText(f"{self.Slider_V_min_Y.value()}")

        self.label_HMax_B.setText(f"{self.Slider_H_max_B.value()}")
        self.label_SMax_B.setText(f"{self.Slider_S_max_B.value()}")
        self.label_VMax_B.setText(f"{self.Slider_V_max_B.value()}")
        self.label_HMin_B.setText(f"{self.Slider_H_min_B.value()}")
        self.label_SMin_B.setText(f"{self.Slider_S_min_B.value()}")
        self.label_VMin_B.setText(f"{self.Slider_V_min_B.value()}")

        self.label_kontras.setText(f"{self.Slider_kontras.value()}")
        self.label_kecerahan.setText(f"{self.Slider_kecerahan.value()}")

        self.displayManualAdjustmentResult()

    def saveToFile(self):
        with open('slider_value.txt', 'w') as file:
            file.write(f"{str(self.Slider_H_max_G.value())},{self.Slider_S_max_G.value()},{self.Slider_V_max_G.value()},{self.Slider_H_min_G.value()},{self.Slider_S_min_G.value()},{self.Slider_V_min_G.value()},{self.Slider_H_max_Y.value()},{self.Slider_S_max_Y.value()},{self.Slider_V_max_Y.value()},{self.Slider_H_min_Y.value()},{self.Slider_S_min_Y.value()},{self.Slider_V_min_Y.value()},{self.Slider_H_max_B.value()},{self.Slider_S_max_B.value()},{self.Slider_V_max_B.value()},{self.Slider_H_min_B.value()},{self.Slider_S_min_B.value()},{self.Slider_V_min_B.value()},{self.Slider_kontras.value()},{self.Slider_kecerahan.value()},")

    # menampilkan masking hasil segmentasi berdasarkamn warna ytang telah ditentukan secara manual
    def displayManualAdjustmentResult(self):
        dd = 1
        if dd == 1:
            image_segmentation2 = cv2.cvtColor(self.masked, cv2.COLOR_BGR2RGB)
            hsv_image = cv2.cvtColor(image_segmentation2, cv2.COLOR_RGB2HSV)

            # Buat mask untuk masing-masing rentang warna
            mask_green = cv2.inRange(
                hsv_image, self.lower_green, self.upper_green)
            mask_yellow = cv2.inRange(
                hsv_image, self.lower_yellow, self.upper_yellow)
            mask_brown = cv2.inRange(
                hsv_image, self.lower_brown, self.upper_brown)

            # Hitung luas daun yang terdeteksi dalam setiap kategori warna
            self.leaf_area_green = np.sum(mask_green == 255)
            self.leaf_area_yellow = np.sum(mask_yellow == 255)
            self.leaf_area_brown = np.sum(mask_brown == 255)
            self.daunmask2 = np.sum(self.daunmask == 255)

            PixelData = [(self.leaf_area_green),
                        (self.leaf_area_yellow), (self.leaf_area_brown)]

            self.mask_combined = cv2.bitwise_or(
            cv2.bitwise_or(mask_green, mask_yellow), mask_brown)

            rgb_image_1 = cv2.cvtColor(self.masked, cv2.COLOR_BGR2RGB)
            rgb_image_2 = cv2.cvtColor(mask_green, cv2.COLOR_BGR2RGB)
            rgb_image_3 = cv2.cvtColor(mask_yellow, cv2.COLOR_BGR2RGB)
            rgb_image_4 = cv2.cvtColor(mask_brown, cv2.COLOR_BGR2RGB)
            self.rgb_image_5 = cv2.cvtColor(self.mask_combined, cv2.COLOR_BGR2RGB)

            h, w, ch = rgb_image_1.shape
            bytes_per_line = ch * w

            convert_to_qt_format_1 = QImage(
                rgb_image_1.data, w, h, bytes_per_line, QImage.Format_RGB888)
            convert_to_qt_format_2 = QImage(
                rgb_image_2.data, w, h, bytes_per_line, QImage.Format_RGB888)
            convert_to_qt_format_3 = QImage(
                rgb_image_3.data, w, h, bytes_per_line, QImage.Format_RGB888)
            convert_to_qt_format_4 = QImage(
                rgb_image_4.data, w, h, bytes_per_line, QImage.Format_RGB888)
            convert_to_qt_format_5 = QImage(
                self.rgb_image_5.data, w, h, bytes_per_line, QImage.Format_RGB888)

            p = convert_to_qt_format_1.scaled(
                720, 640, aspectRatioMode=True)
            p2 = convert_to_qt_format_2.scaled(
                320, 240, aspectRatioMode=True)
            p3 = convert_to_qt_format_3.scaled(
                320, 240, aspectRatioMode=True)
            p4 = convert_to_qt_format_4.scaled(
                320, 240, aspectRatioMode=True)
            self.p5 = convert_to_qt_format_5.scaled(
                320, 240, aspectRatioMode=True)
                        
            self.disp_main_2.setPixmap(QPixmap.fromImage(p2))
            self.disp_main_3.setPixmap(QPixmap.fromImage(p3))
            self.disp_main_4.setPixmap(QPixmap.fromImage(p4))
            self.disp_main_5.setPixmap(QPixmap.fromImage(self.p5))
    
    def on_value_change(self):
        self.lower_green = np.array([self.Slider_H_min_G.value(
        ), self.Slider_S_min_G.value(), self.Slider_V_min_G.value()])
        self.upper_green = np.array([self.Slider_H_max_G.value(
        ), self.Slider_S_max_G.value(), self.Slider_V_max_G.value()])

        self.lower_yellow = np.array([self.Slider_H_min_Y.value(
        ), self.Slider_S_min_Y.value(), self.Slider_V_min_Y.value()])
        self.upper_yellow = np.array([self.Slider_H_max_Y.value(
        ), self.Slider_S_max_Y.value(), self.Slider_V_max_Y.value()])

        self.lower_brown = np.array([self.Slider_H_min_B.value(
        ), self.Slider_S_min_B.value(), self.Slider_V_min_B.value()])
        self.upper_brown = np.array([self.Slider_H_max_B.value(
        ), self.Slider_S_max_B.value(), self.Slider_V_max_B.value()])

        self.update_hsv(self.lower_green, self.upper_green, self.lower_yellow,
                        self.upper_yellow, self.lower_brown, self.upper_brown, self.Slider_kontras.value(),self.Slider_kecerahan.value())

    def save_daun(self):
        if self.freeze:
            self.save_data_daun()
            self.DateTime = QDateTime.currentDateTime()
            self.textEdit.append(f"{self.DateTime.toString('d MMMM yy hh:mm:ss')}: Data Successfully Saved")

    def save_data_daun(self):

        self.data_daun = np.array([(self.DateTime.toString('d MMMM yy hh:mm:ss')), 
                                   (str(self.panjangobjek)+" mm"), (str(self.lebarobjek)+" mm"), (str(self.luasobjek)+" mm"), 
                                   (str(self.leaf_area_green)+" pixel"), (str(self.leaf_area_yellow)+" pixel"), (str(self.leaf_area_brown)+" pixel"), 
                                   (str(self.presentase_green) + " %"), (str(self.presentase_yellow) + " %"), (str(self.presentase_brown) + " %")])
        print("Ini Data Daun : ", self.data_daun)
        self.add_data_to_excel(self.data_daun)


    def add_data_to_excel(self, person_data):
        if not os.path.exists(self.file_name):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            header = ["No", "Tanggal","Panjang Daun", "Lebar Daun","Luas Daun", 
                      "Luas Hijau", "Luas Kuning","Luas Coklat",
                      "Presentase Warna Hijau", "Presentase Warna Kuning", "Presentase Warna Coklat"]
            sheet.append(header)
            current_row = 1 
        else:
            workbook = openpyxl.load_workbook(self.file_name)
            sheet = workbook.active
            current_row = sheet.max_row
        sheet.append([current_row] + person_data.tolist())

        last_row = sheet.max_row

        for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=last_row):
            for cell in col:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

        workbook.save(self.file_name)

    #menampilkan nilai panjang, lebar, dan luas daun ke GUI
    def displayData(self):
        # Display data for the first detected object if available
        if self.objectdetected > 0:
            self.display_panjang0.setText(f"{self.panjangobjek[0]} mm")
            self.display_lebar0.setText(f"{self.lebarobjek[0]} mm")
            self.display_luas0.setText(f"{self.luasobjek[0]} mm")
        else:
            self.display_panjang0.setText("0 mm")
            self.display_lebar0.setText("0 mm")
            self.display_luas0.setText("0 mm")

        # Display data for the second detected object if available
        if self.objectdetected > 1:
            self.display_panjang1.setText(f"{self.panjangobjek[1]} mm")
            self.display_lebar1.setText(f"{self.lebarobjek[1]} mm")
            self.display_luas1.setText(f"{self.luasobjek[1]} mm")
        else:
            self.display_panjang1.setText("0 mm")
            self.display_lebar1.setText("0 mm")
            self.display_luas1.setText("0 mm")

        # Display data for the third detected object if available
        if self.objectdetected > 2:
            self.display_panjang2.setText(f"{self.panjangobjek[2]} mm")
            self.display_lebar2.setText(f"{self.lebarobjek[2]} mm")
            self.display_luas2.setText(f"{self.luasobjek[2]} mm")
        else:
            self.display_panjang2.setText("0 mm")
            self.display_lebar2.setText("0 mm")
            self.display_luas2.setText("0 mm")


    def klik_apply(self):
        self.DateTime = QDateTime.currentDateTime()
        self.display_clock.display(self.DateTime.toString('hh:mm:ss'))
        #masked2 = cv2.bitwise_and(self.image2, self.image2, mask=self.p5)
        #pixmap = QPixmap(self.rgb_image_5)
        #print(self.image2.shape)
        #print(self.mask_combined.shape)
        self.masked2 = cv2.bitwise_and(self.image2, self.image2, mask=self.mask_combined)
        #print(self.masked2.shape)
        self.masked3 = cv2.cvtColor(self.masked2, cv2.COLOR_BGR2RGB)
        h, w, ch = self.masked3.shape
        bytes_per_line = ch * w
        convert_to_qt_format_x = QImage(
            self.masked3.data, w, h, bytes_per_line, QImage.Format_RGB888)
        self.px = convert_to_qt_format_x.scaled(
            740, 450, aspectRatioMode=True)                        
        self.disp_main_1.setPixmap(QPixmap.fromImage(self.px))

    def klik_restore(self):
        self.DateTime = QDateTime.currentDateTime()
        self.display_clock.display(self.DateTime.toString('hh:mm:ss'))
        pixmap = QPixmap(self.filename)
        pixmap = pixmap.scaled(740, 450)
        self.disp_main_1.setPixmap(pixmap)
        
    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "UGM - Miconos Leaf Area Meter"))
        self.groupBox_3.setTitle(_translate("MainWindow", "   Leaf Area Meter - Detection"))
        self.groupBox_2.setTitle(_translate("MainWindow", "Choose Camera"))
        self.groupBox_4.setTitle(_translate("MainWindow", "Main Camera"))
        self.disp_main_1.setText(_translate("MainWindow", "Main Source"))
        self.groupBox_9.setTitle(_translate("MainWindow", "Segmentation"))
        self.disp_main_2.setText(_translate("MainWindow", "Display Mask Green"))
        self.disp_main_3.setText(_translate("MainWindow", "Display Mask Yellow"))
        self.disp_main_4.setText(_translate("MainWindow", "Display Mask Brown"))
        self.disp_main_5.setText(_translate("MainWindow", "Display Mask Green/Yellow/Brown"))
        self.label.setText(_translate("MainWindow", "GREEN"))
        self.label_10.setText(_translate("MainWindow", "YELLOW"))
        self.label_12.setText(_translate("MainWindow", "BROWN"))
        self.label_13.setText(_translate("MainWindow", "GREEN - YELLOW - BROWN"))
        self.ckb_data_2.setText(_translate("MainWindow", "Tampilkan"))
        self.btn_apply.setToolTip(_translate("MainWindow", "Apply this mask to the image"))
        self.btn_apply.setText(_translate("MainWindow", "APPLY"))
        self.btn_restore.setToolTip(_translate("MainWindow", "Restore original segmentation"))
        self.btn_restore.setText(_translate("MainWindow", "RESTORE"))
        self.groupBox_13.setTitle(_translate("MainWindow", "HSV Setting"))
        self.groupBox_6.setTitle(_translate("MainWindow", "HSV Setting Green"))
        self.label_4.setText(_translate("MainWindow", "HSV Max"))
        self.label_6.setText(_translate("MainWindow", "HSV Min"))
        self.label_7.setText(_translate("MainWindow", "H"))
        self.label_SMax_G.setText(_translate("MainWindow", "0"))
        self.label_VMax_G.setText(_translate("MainWindow", "0"))
        self.label_HMax_G.setText(_translate("MainWindow", "0"))
        self.label_5.setText(_translate("MainWindow", "H"))
        self.label_8.setText(_translate("MainWindow", "S"))
        self.label_9.setText(_translate("MainWindow", "V"))
        self.label_11.setText(_translate("MainWindow", "S"))
        self.label_SMin_G.setText(_translate("MainWindow", "0"))
        self.label_14.setText(_translate("MainWindow", "V"))
        self.label_VMin_G.setText(_translate("MainWindow", "0"))
        self.label_HMin_G.setText(_translate("MainWindow", "0"))
        self.groupBox_7.setTitle(_translate("MainWindow", "HSV Setting Yellow"))
        self.label_19.setText(_translate("MainWindow", "HSV Min"))
        self.label_20.setText(_translate("MainWindow", "H"))
        self.label_SMax_Y.setText(_translate("MainWindow", "0"))
        self.label_VMax_Y.setText(_translate("MainWindow", "0"))
        self.label_HMax_Y.setText(_translate("MainWindow", "0"))
        self.label_24.setText(_translate("MainWindow", "H"))
        self.label_25.setText(_translate("MainWindow", "S"))
        self.label_26.setText(_translate("MainWindow", "V"))
        self.label_27.setText(_translate("MainWindow", "S"))
        self.label_SMin_Y.setText(_translate("MainWindow", "0"))
        self.label_29.setText(_translate("MainWindow", "V"))
        self.label_VMin_Y.setText(_translate("MainWindow", "0"))
        self.label_HMin_Y.setText(_translate("MainWindow", "0"))
        self.label_18.setText(_translate("MainWindow", "HSV Max"))
        self.groupBox_8.setTitle(_translate("MainWindow", "HSV Setting Brown"))
        self.label_32.setText(_translate("MainWindow", "HSV Max"))
        self.label_33.setText(_translate("MainWindow", "HSV Min"))
        self.label_34.setText(_translate("MainWindow", "H"))
        self.label_SMax_B.setText(_translate("MainWindow", "0"))
        self.label_VMax_B.setText(_translate("MainWindow", "0"))
        self.label_HMax_B.setText(_translate("MainWindow", "0"))
        self.label_38.setText(_translate("MainWindow", "H"))
        self.label_39.setText(_translate("MainWindow", "S"))
        self.label_40.setText(_translate("MainWindow", "V"))
        self.label_41.setText(_translate("MainWindow", "S"))
        self.label_SMin_B.setText(_translate("MainWindow", "0"))
        self.label_43.setText(_translate("MainWindow", "V"))
        self.label_VMin_B.setText(_translate("MainWindow", "0"))
        self.label_HMin_B.setText(_translate("MainWindow", "0"))
        self.groupBox_11.setTitle(_translate("MainWindow", "Brightness Contrast"))
        self.label_kontras.setText(_translate("MainWindow", "0"))
        self.label_51.setText(_translate("MainWindow", "Brightness "))
        self.label_kecerahan.setText(_translate("MainWindow", "0"))
        self.label_52.setText(_translate("MainWindow", "Contrast     "))
        self.label_35.setText(_translate("MainWindow", "Brightness and Contrast"))
        self.groupBox_14.setTitle(_translate("MainWindow", "Display"))
        self.label_85.setText(_translate("MainWindow", "Panjang Daun"))
        self.display_luas1.setText(_translate("MainWindow", "0.0"))
        self.display_lebar1.setText(_translate("MainWindow", "0.0"))
        self.label_86.setText(_translate("MainWindow", "Lebar Daun"))
        self.label_87.setText(_translate("MainWindow", "Luas Area"))
        self.display_panjang1.setText(_translate("MainWindow", "0.0"))
        self.label_103.setText(_translate("MainWindow", "Panjang Daun"))
        self.display_luas2.setText(_translate("MainWindow", "0.0"))
        self.display_lebar2.setText(_translate("MainWindow", "0.0"))
        self.label_104.setText(_translate("MainWindow", "Lebar Daun"))
        self.label_105.setText(_translate("MainWindow", "Luas Area"))
        self.display_panjang2.setText(_translate("MainWindow", "0.0"))
        self.label_112.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-weight:600;\">Daun 1</span></p></body></html>"))
        self.label_113.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-weight:600;\">Daun 2</span></p></body></html>"))
        self.label_114.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-weight:600;\">Daun 3</span></p></body></html>"))
        self.label_17.setText(_translate("MainWindow", "Lebar Daun"))
        self.display_panjang0.setText(_translate("MainWindow", "0.0"))
        self.label_36.setText(_translate("MainWindow", "Luas Area"))
        self.display_lebar0.setText(_translate("MainWindow", "0.0"))
        self.label_15.setText(_translate("MainWindow", "Panjang Daun"))
        self.display_luas0.setText(_translate("MainWindow", "0.0"))
        # self.display_presentase_yellow.setText(_translate("MainWindow", "0.0"))
        # self.label_44.setText(_translate("MainWindow", "Presentase Coklat"))
        # self.label_45.setText(_translate("MainWindow", "Presentase Kuning"))
        # self.display_presentase_green.setText(_translate("MainWindow", "0.0"))
        # self.display_presentase_brown.setText(_translate("MainWindow", "0.0"))
        self.btn_save.setToolTip(_translate("MainWindow", "Save measurment data"))
        self.btn_save.setText(_translate("MainWindow", "SAVE"))
        self.groupBox_5.setTitle(_translate("MainWindow", "Information"))
        self.groupBox_12.setTitle(_translate("MainWindow", "Computer Information"))
        self.label_16.setText(_translate("MainWindow", "CPU"))
        self.Qlabel_cpu.setText(_translate("MainWindow", "0.0"))
        self.Qlabel_ram.setText(_translate("MainWindow", "0.0"))
        self.label_28.setText(_translate("MainWindow", "RAM"))
        self.Qlabel_fps.setText(_translate("MainWindow", "0.0"))
        self.label_21.setText(_translate("MainWindow", "FPS"))
        self.groupBox_10.setTitle(_translate("MainWindow", "Pushbutton"))
        self.btn_start_1.setToolTip(_translate("MainWindow", "Preview camera"))
        self.btn_start_1.setText(_translate("MainWindow", "OPEN CAMERA"))
        self.btn_start_2.setToolTip(_translate("MainWindow", "Capture image"))
        self.btn_start_2.setText(_translate("MainWindow", "TAKE PICTURE"))
        self.btn_start_3.setToolTip(_translate("MainWindow", "Detect Leaf"))
        self.btn_start_3.setText(_translate("MainWindow", "ANALYSIS"))
        self.btn_not_ai.setText(_translate("MainWindow", "NOT USE AI"))
        self.btn_calibration.setText(_translate("MainWindow", "Calibration"))

'''    def iniPP(self):
        cpu = pyqtSignal(float)
        ram = pyqtSignal(tuple)
        temp = pyqtSignal(float)
        ProsesJalan = True
        while ProsesJalan:
            cpu = sysinfo.getCPU()
            ram = sysinfo.getRAM()
            self.getCPU_usage(cpu)
            self.getRAM_usage(ram)

    def getCPU_usage(self, cpu):
        self.Qlabel_cpu.setText(str(cpu) + " %")
        if cpu > 15:
            self.Qlabel_cpu.setStyleSheet("color: rgb(23, 63, 95);")
        if cpu > 25:
            self.Qlabel_cpu.setStyleSheet("color: rgb(32, 99, 155);")
        if cpu > 45:
            self.Qlabel_cpu.setStyleSheet("color: rgb(60, 174, 163);")
        if cpu > 65:
            self.Qlabel_cpu.setStyleSheet("color: rgb(246, 213, 92);")
        if cpu > 85:
            self.Qlabel_cpu.setStyleSheet("color: rgb(237, 85, 59);")

    def getRAM_usage(self, ram):
        self.Qlabel_ram.setText(str(ram[2]) + " %")
        if ram[2] > 15:
            self.Qlabel_ram.setStyleSheet("color: rgb(23, 63, 95);")
        if ram[2] > 25:
            self.Qlabel_ram.setStyleSheet("color: rgb(32, 99, 155);")
        if ram[2] > 45:
            self.Qlabel_ram.setStyleSheet("color: rgb(60, 174, 163);")
        if ram[2] > 65:
            self.Qlabel_ram.setStyleSheet("color: rgb(246, 213, 92);")
        if ram[2] > 85:
            self.Qlabel_ram.setStyleSheet("color: rgb(237, 85, 59);")

    def getTemp_usage(self, temp):
        self.Qlabel_temp.setText(str(temp) + " *C")
        if temp > 30:
            self.Qlabel_temp.setStyleSheet("color: rgb(23, 63, 95);")
        if temp > 35:
            self.Qlabel_temp.setStyleSheet("color: rgb(60, 174, 155);")
        if temp > 40:
            self.Qlabel_temp.setStyleSheet("color: rgb(246,213, 92);")
        if temp > 45:
            self.Qlabel_temp.setStyleSheet("color: rgb(237, 85, 59);")
        if temp > 50:
            self.Qlabel_temp.setStyleSheet("color: rgb(255, 0, 0);")

    def get_FPS(self, fps):
        # print("fps di sini : ", str(fps))
        self.Qlabel_fps.setText(str(fps))
        if fps > 5:
            self.Qlabel_fps.setStyleSheet("color: rgb(237, 85, 59);")
        if fps > 15:
            self.Qlabel_fps.setStyleSheet("color: rgb(60, 174, 155);")
        if fps > 25:
            self.Qlabel_fps.setStyleSheet("color: rgb(85, 170, 255);")
        if fps > 35:
            self.Qlabel_fps.setStyleSheet("color: rgb(23, 63, 95);")'''

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
